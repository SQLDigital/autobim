import copy
from openpyxl import Workbook, load_workbook
from tempfile import NamedTemporaryFile
from django.db.models import F, Count
from django.core.paginator import Paginator
from django.shortcuts import render, redirect, HttpResponse
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import View, ListView, DetailView, TemplateView, FormView
from django.views.generic.edit import UpdateView, DeleteView, CreateView, FormView
from django.contrib.auth.hashers import make_password
from django.urls import reverse_lazy
from django.http import JsonResponse
from django.conf import settings
from bootstrap_modal_forms.generic import BSModalCreateView, BSModalUpdateView, BSModalDeleteView, BSModalFormView

from users.models import User
from design_health.views import AjaxableResponseMixin

from .models import Discipline, Role, Project, ProjectRole, ProjectTask, Task, TaskType, TaskLeader, Company, TaskMember, DisciplineCategory, ProjectScale

from . import forms


def is_super_user(user):
    return user.is_superuser

def is_task_leader(user):
    return user.role == 'task_leader'

def is_team_member(user):
    return user.role == 'team_member'

class SuperuserOnlyMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_superuser


@login_required
@user_passes_test(is_super_user)
def role_list_view(request):
    if request.method == 'POST':
        form = forms.RoleForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(request.path)

    form = forms.RoleForm()
    role_list = Role.objects.all()
    paginator = Paginator(role_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        "form": form,
        "page_obj": page_obj,
    }

    return render(request, 'tidp/roles.html', context)


class RoleUpdateView(LoginRequiredMixin, UserPassesTestMixin, BSModalUpdateView):
    model = Role
    form_class = forms.RoleFormPopup
    success_message = 'Role successfully updated.'
    success_url = reverse_lazy('tidp:roles')
    template_name = 'tidp/popup_update_form.html'

    def test_func(self):
      return self.request.user.is_superuser



class RoleDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Role
    template_name = "tidp/delete_form.html"
    success_url = reverse_lazy('tidp:roles')
    success_message = 'Role succesfully deleted.'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = "Role Delete"
        return context

    def test_func(self):
      return self.request.user.is_superuser

    def delete(self, request, *args, **kwargs):
        messages.success(self.request, self.success_message)
        return super().delete(request, *args, **kwargs)


class CompanyListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Company
    template_name = "tidp/company_list.html"
    paginate_by = 10

    def test_func(self):
        return self.request.user.is_superuser


class CompanyCreateView(LoginRequiredMixin, BSModalCreateView):
    model = Company
    form_class = forms.CompanyForm
    template_name = 'tidp/popup_create_form.html'
    success_message = 'New company added'
    success_url = reverse_lazy('tidp:company-list')

    def get_success_url(self):
        if self.request.POST.get('next'):
            return self.request.POST.get('next')
        return self.success_url


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add new company'
        return context



class CompanyUpdateView(LoginRequiredMixin, BSModalUpdateView):
    model = Company
    form_class = forms.CompanyForm
    template_name = 'tidp/popup_update_form.html'
    success_message = 'Company updated successfully'
    success_url = reverse_lazy('tidp:company-list')

    def get_success_url(self):
        if self.request.POST.get('next'):
            return self.request.POST.get('next')
        return self.success_url


class CompanyDeleteView(LoginRequiredMixin, UserPassesTestMixin, BSModalDeleteView):
    model = Company
    template_name = 'tidp/delete_form.html'
    success_url = reverse_lazy('tidp:company-list')
    success_message = 'Company Deleted'

    def test_func(self):
        return self.request.user.is_superuser


class ManageProjects(LoginRequiredMixin, ListView):
    """ Manage all projects """
    model = Project
    template_name = 'tidp/project/project_list.html'

    def get_queryset(self):
        queryset = Project.objects.none()
        params = { x:y for x,y in self.request.GET.items() if y}
        if params:
            queryset = super().get_queryset().filter(**params)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        params = { x:y for x,y in self.request.GET.items() if y}
        context['searchform'] = forms.SearchProject(initial=params)
        return context


class ProjectCreateView(LoginRequiredMixin, BSModalCreateView):
    model = Project
    form_class = forms.ProjectForm
    success_message = 'New project added'
    template_name = "tidp/project/project_form.html"

    def form_valid(self, form):
        form.instance.author = self.request.user
        form.save()
        return super().form_valid(form)


class ProjectDetailView(SuperuserOnlyMixin, View):
    """ Manage a single project """
    template_name = 'tidp/project/project_detail.html'

    def get(self, request, *args, **kwargs):
        project = Project.objects.get(pk=kwargs['pk'])
        task_leaders = project.taskleader_set.all()
        task_members = project.taskmember_set.all()
        roles = project.projectrole_set.all()
        """ tasks = ProjectTask.objects.filter(project=project).order_by('role', 'company')
        grouped = {}
        for obj in tasks:
            grouped.setdefault((obj.role, obj.company), []).append(obj)

        items = []
        for key, group in grouped.items():
            items.append((key[0], key[1], any([s.is_active for s in group]), group)) """


        """roles = []
        for task in projecttasks.values('role').distinct():
            roles.append(
                (
                    task['role'],
                    Role.objects.values_list('name', flat=True).get(pk=task['role']),
                    projecttasks.filter(role=task['role']),
                )
            ) """

        context = {
            "project": project,
            "roles": roles,
            "task_leaders": task_leaders,
            "task_members": task_members,

        }
        return render(request, self.template_name, context)


class ProjectDetailLeaderView(LoginRequiredMixin, View):
    """ Manage a single project for project leader """
    template_name = 'tidp/project/project_detail_leader.html'

    def get(self, request, *args, **kwargs):
        project = Project.objects.get(pk=kwargs['pk'])
        task_members = project.taskmember_set.all()
        taskleader = TaskLeader.objects.get(project=project, user=request.user)
        roles = project.projectrole_set.filter(role__in=taskleader.roles.all())

        context = {
            "project": project,
            "roles": roles,
            "task_members": task_members,
        }
        return render(request, self.template_name, context)


class ProjectUpdateView(LoginRequiredMixin, UserPassesTestMixin, BSModalUpdateView):
    model = Project
    form_class = forms.ProjectForm
    template_name = "tidp/popup_update_form.html"
    success_message = 'Project successfully updated.'

    def test_func(self):
        return self.request.user.is_superuser


class ProjectDeleteView(LoginRequiredMixin, BSModalDeleteView):
    model = Project
    success_url = reverse_lazy('tidp:projects')
    template_name = "tidp/delete_form.html"
    success_message = 'Project deleted.'


class ProjectSubmitToLeadView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        data = request.POST['projectrole']
        projectrole = ProjectRole.objects.filter(pk=data).update(status="pending")
        return JsonResponse({"message": "Success"})


class RoleStatusUpdateView(LoginRequiredMixin, View):
    template_name = 'tidp/popup_update_form.html'
    form_class = forms.RoleStatusForm

    def get(self, request, *args, **kwargs):
        projectrole = ProjectRole.objects.get(pk=kwargs['projectrole_id'])
        form = self.form_class(instance=projectrole)
        context = {
            "form": form,
            "title": "Add new form",
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        projectrole = ProjectRole.objects.get(pk=kwargs['projectrole_id'])
        form = self.form_class(request.POST, instance=projectrole)
        if form.is_valid():
            projectrole.status = form.cleaned_data['status']
            projectrole.comment = form.cleaned_data['comment']
            projectrole.save()
            messages.success(request, 'Project Role successfully updated.')
            return redirect('tidp:project-detail', projectrole.project.id)

        context = {
            "form": form,
            "title": "Add new form",
        }
        return render(request, self.template_name, context)


class DisciplineListView(LoginRequiredMixin, ListView):
    model = Discipline
    template_name = 'tidp/discipline_list.html'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('name')
        if search:
            queryset = queryset.filter(name__icontains=search)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search'] = forms.DisciplineSearchForm(initial={"name":self.request.GET.get('name')})
        return context


class DisciplineCreateView(LoginRequiredMixin, BSModalCreateView):
    model = Discipline
    form_class = forms.DisciplineForm
    template_name = "tidp/popup_create_form.html"
    success_message = 'Discipline successfully added.'
    success_url = reverse_lazy(('tidp:project-discipline'))

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add new discipline'
        return context


class DisciplineUpdateView(LoginRequiredMixin, BSModalUpdateView):
    model = Discipline
    form_class = forms.DisciplineForm
    template_name = "tidp/popup_update_form.html"
    success_url = reverse_lazy('tidp:project-discipline')
    success_message = 'Discipline successfully updated.'


class DisciplineDeleteView(LoginRequiredMixin, BSModalDeleteView):
    model = Discipline
    success_url = reverse_lazy('tidp:project-discipline')
    template_name = "tidp/delete_form.html"
    success_message = 'Discipline successfully deleted.'



class DisciplineCategoryListView(LoginRequiredMixin, ListView):
    model = DisciplineCategory
    template_name = 'tidp/discipline_category.html'
    context_object_name = 'disciplines'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('name')
        if search:
            queryset = queryset.filter(category__icontains=search)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search'] = forms.DisciplineSearchForm(initial={"name":self.request.GET.get('name')})
        return context


class DisciplineCategoryCreateView(LoginRequiredMixin, BSModalCreateView):
    models = DisciplineCategory
    form_class = forms.DisciplineCategoryForm
    success_url = reverse_lazy('tidp:project-discipline-categories')
    success_message = 'Discipline category successfully added.'
    template_name = 'tidp/popup_create_form.html'


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add new Discipline Category'
        return context



class DisciplineCategoryUpdateView(LoginRequiredMixin, BSModalUpdateView):
    model = DisciplineCategory
    form_class = forms.DisciplineCategoryForm
    template_name = "tidp/popup_update_form.html"
    success_url = reverse_lazy('tidp:project-discipline-categories')
    success_message = 'Discipline Category successfully updated.'


class DisciplineCategoryDeleteView(LoginRequiredMixin, BSModalDeleteView):
    model = DisciplineCategory
    success_url = reverse_lazy('tidp:project-discipline-categories')
    template_name = "tidp/delete_form.html"
    success_message = 'Discipline Category delete.'



class TaskListView(LoginRequiredMixin, ListView):
    model = Task
    template_name = 'tidp/tasks/task_list.html'
    context_object_name = 'tasks'
    paginate_by = 20

    def get_queryset(self):
        queryset = super().get_queryset()
        params = {x:y for x,y in self.request.GET.items() if y}
        if params:
            queryset = queryset.filter(**params)
            print(queryset)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        initials = {x:y for x,y in self.request.GET.items() if y}
        context['search'] = forms.TaskSearchForm(initial=initials)
        return context


class TaskCreateView(LoginRequiredMixin, BSModalCreateView):
    model = Task
    form_class = forms.TaskForm
    template_name = "tidp/popup_create_form.html"
    success_url = reverse_lazy('tidp:tasks')
    success_message = 'Task successfully added.'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add new task'
        return context

    def get_form_kwargs(self, **kwargs):
        form_kwargs = super().get_form_kwargs(**kwargs)
        form_kwargs["prefix"] = 'task'
        return form_kwargs


class TaskUpdateView(LoginRequiredMixin, BSModalUpdateView):
    model = Task
    form_class = forms.TaskForm
    template_name = "tidp/popup_update_form.html"
    success_url = reverse_lazy('tidp:tasks')
    success_message = 'Task successfully updated.'


class TaskDeleteView(LoginRequiredMixin, BSModalDeleteView):
    model = Task
    success_url = reverse_lazy('tidp:tasks')
    template_name = "tidp/delete_form.html"
    success_message = 'Task succesfully deleted.'


class TaskTypeView(LoginRequiredMixin, ListView):
    model = TaskType
    template_name = 'tidp/tasks/tasktype.html'
    paginate_by = 20
    context_object_name = 'tasktypes'
    ordering = ['title']


class TaskTypeCreateView(LoginRequiredMixin, BSModalCreateView):
    model = TaskType
    form_class = forms.TaskTypeForm
    template_name = 'tidp/popup_create_form.html'
    success_message = 'New task type successfully added.'
    success_url = reverse_lazy('tidp:tasktypes')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add new taskType'
        return context


class TaskTypeUpdateView(LoginRequiredMixin, BSModalUpdateView):
    model = TaskType
    form_class = forms.TaskTypeForm
    template_name = "tidp/popup_update_form.html"
    success_url = reverse_lazy('tidp:tasktypes')
    success_message = 'Task type successfully updated.'


class TaskTypeDeleteView(LoginRequiredMixin, BSModalDeleteView):
    model = TaskType
    success_url = reverse_lazy('tidp:tasktypes')
    template_name = "tidp/delete_form.html"
    success_message = 'Task type succesfully deleted.'



class TaskLeaderCreateView(LoginRequiredMixin, BSModalCreateView):
    model = TaskLeader
    form_class = forms.TaskLeaderForm
    template_name = "tidp/popup_create_form.html"
    success_message = 'Task leader successfully added.'

    def form_valid(self, form):
        user, iscreated = User.objects.get_or_create(
            email=form.cleaned_data['contact_email'],
            defaults={"password":make_password('1a@p8xazi')}
        )
        form.instance.user = user
        form.instance.project_id = self.kwargs['project_id']
        exist = TaskLeader.objects.filter(user=user, project_id=self.kwargs['project_id']).exists()
        if exist:
            form.add_error('__all__', 'Task leader already exists!')
            return super().form_invalid(form)
        return super().form_valid(form)

    def get_success_url(self):
        obj = self.object
        return reverse_lazy('tidp:project-detail', kwargs={'pk': obj.project.id})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add new task leader'
        return context


class TaskLeaderUpdateView(LoginRequiredMixin, BSModalUpdateView):
    model = TaskLeader
    form_class = forms.TaskLeaderForm
    template_name = "tidp/popup_update_form.html"
    success_message = 'Task leader successfully updated.'

    def get_success_url(self):
        obj = self.object
        return reverse_lazy('tidp:project-detail', kwargs={"pk":obj.project.id})


class TaskLeaderDeleteView(LoginRequiredMixin, BSModalDeleteView):
    model = TaskLeader
    template_name = "tidp/delete_form.html"
    success_message = 'Task leader succesfull deleted.'

    def get_success_url(self):
        obj = self.object
        return reverse_lazy('tidp:project-detail', kwargs={"pk":obj.project.id})


class TaskMemberCreateView(LoginRequiredMixin, BSModalCreateView):
    model = TaskMember
    form_class = forms.TaskMemberForm
    template_name = "tidp/popup_create_form.html"
    success_message = 'A new task member has been added successfully.'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs.update({'project': self.kwargs['project_id']})
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add new task member'
        return context

    def get_success_url(self):
        obj = self.object
        if self.request.user.is_superuser:
            url = 'tidp:project-detail'
        else:
            url = 'tidp:project-detail-leader'
        return reverse_lazy(url, args=[obj.project.id])

    def form_valid(self, form):
        user, iscreated = User.objects.get_or_create(
            email=form.cleaned_data['contact_email'],
            defaults={"password":make_password('1a@p8xazi')}
        )
        form.instance.user = user
        form.instance.project_id = self.kwargs['project_id']

        exist = TaskMember.objects.filter(user=user, project_id=self.kwargs['project_id']).exists()
        if exist:
            form.add_error('__all__', 'Task member already exists!')
            return super().form_invalid(form)

        try:
            leader = TaskLeader.objects.get(user=self.request.user, project=self.kwargs['project_id'])
            form.instance.company = leader.company
        except TaskLeader.DoesNotExist:
            form.add_error('__all__', 'Task leader does not exists!')
            return super().form_invalid(form)

        return super().form_valid(form)


class TaskMemberUpdateView(LoginRequiredMixin, BSModalUpdateView):
    model = TaskMember
    form_class = forms.TaskMemberForm
    template_name = "tidp/popup_update_form.html"
    success_message = 'Task member successfully updated.'

    def get_success_url(self):
        obj = self.object
        if self.request.user.is_superuser:
            url = 'tidp:project-detail'
        else:
            url = 'tidp:project-detail-leader'
        return reverse_lazy(url, kwargs={"pk":obj.project.id})


class TaskMemberDeleteView(LoginRequiredMixin, BSModalDeleteView):
    model = TaskMember
    template_name = "tidp/delete_form.html"
    success_message = 'Task member successfully deleted.'

    def get_success_url(self):
        obj = self.object
        if self.request.user.is_superuser:
            url = 'tidp:project-detail'
        else:
            url = 'tidp:project-detail-leader'
        return reverse_lazy(url, kwargs={"pk":obj.project.id})



class ProjectTaskCreateView(LoginRequiredMixin, View):
    form_class = forms.ProjectTaskCreateForm
    template_name = "tidp/project/projecttask_create.html"
    success_message = 'New tasks has been added successfully.'

    def get(self, request, *args, **kwargs):
        context = {
            "form" : self.form_class,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            if form.cleaned_data['clone'] == 'yes':
                request.session['tidp_stuff'] = {
                    "description": form.cleaned_data['description'],
                    "uniclass_code": form.cleaned_data['uniclass_code'],
                    "task_type": form.cleaned_data['task_type'].id,
                    "project_id": kwargs['project_id'],
                    "discipline_id": kwargs['discipline'],
                    "discipline_category_id": kwargs['discipline_category'],
                    "role_id": kwargs['role'],
                    "company_id": kwargs['company'],
                    "clone": form.cleaned_data['clone'],
                    "clone_count": form.cleaned_data['clone_count'],
                }
                return redirect('tidp:projecttask-createmultiple')
            else:
                ProjectTask.objects.create(
                    discipline_id=kwargs['discipline'],
                    project_role_id=kwargs['role'],
                    discipline_category_id=kwargs['discipline_category'],
                    uniclass_code= form.cleaned_data['uniclass_code'],
                    description= form.cleaned_data['description'],
                    task_type= form.cleaned_data['task_type'],
                )
                ProjectRole.objects.filter(project=kwargs['project_id'],role=kwargs['role'], company=kwargs['company']).update(is_active=True)
                messages.success(request, "Project task succesfully added.")
                if request.user.is_superuser:
                    url = 'tidp:project-detail'
                else:
                    url = 'tidp:project-detail-leader'
                return redirect(url, kwargs['project_id'])


class ProjectTaskDuplicateView(LoginRequiredMixin, View):
    form_class = forms.DuplicateCreateForm
    template_name = "tidp/multiple_task_create.html"
    success_message = 'New tasks has been added successfully.'

    def dispatch(self, request, *args, **kwargs):
        self.items = self.request.session.get('tidp_stuff')
        return super().dispatch(request, *args, **kwargs)

    def get_initials(self):
        initials = [
            {
            "description":f"{self.items['description']}_{x}",
            "uniclass_code": self.items['uniclass_code'],
            "task_type":self.items['task_type']
            } for x in range(1, self.items['clone_count']+1)
        ]
        return initials

    def get(self, request, *args, **kwargs):
        initials = self.get_initials()
        formset = self.form_class(queryset=ProjectTask.objects.none(), initial=initials)

        formset.min_num = self.items['clone_count']
        formset.max_num = self.items['clone_count']

        context = {
            "formset" : formset,
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        formset = self.form_class(request.POST)
        if formset.is_valid():
            forms = formset.save(commit=False)
            items = self.items
            for form in forms:
                form.discipline_id = items['discipline_id']
                form.discipline_category_id = items['discipline_category_id']
                form.project_role_id = items['role_id']
                form.save()

            project_id = items['project_id']
            ProjectRole.objects.filter(project_id=project_id,role=items['role_id'], company=items['company_id']).update(is_active=True)

            del items

            messages.success(request, self.success_message)
            if request.user.is_superuser:
                url = 'tidp:project-detail'
            else:
                url = 'tidp:project-detail-leader'
            return redirect(url, project_id)

        formset.min_num = items['clone']
        formset.max_num = items['clone']

        context = {
            "formset" : formset,
        }
        return render(request, self.template_name, context)


class ProjectTaskUpdateView(LoginRequiredMixin, BSModalUpdateView):
    model = ProjectTask
    form_class = forms.ProjectTaskForm
    template_name = "tidp/popup_update_form.html"
    success_message = 'Project task successfully updated.'

    def get_success_url(self):
        obj = self.object
        if self.request.user.is_superuser:
            url = 'tidp:project-detail'
        else:
            url = 'tidp:project-detail-leader'
        return reverse_lazy(url, kwargs={"pk":obj.project_role.project.id})

    def form_valid(self, form):
        ProjectRole.objects.filter(pk=form.instance.project_role.id).update(is_active=True)

        return super().form_valid(form)


class ProjectTaskDetailView(LoginRequiredMixin, DetailView):
    model = ProjectTask
    template_name = 'tidp/tasks/task_detail.html'
    context_object_name = 'task'


class ProjectTaskDeleteView(LoginRequiredMixin, DeleteView):
    model = ProjectTask
    template_name = "tidp/delete_form.html"
    success_message = 'Project Task successfully deleted.'

    def get_success_url(self):
        obj = self.object
        if self.request.user.is_superuser:
            url = 'tidp:project-detail'
        else:
            url = 'tidp:project-detail-leader'
        return reverse_lazy(url, kwargs={"pk":obj.project_role.project.id})

    def delete(self, request, *args, **kwargs):
        response = super(ProjectTaskDeleteView, self).delete(request, *args, **kwargs)
        obj = self.object
        ProjectRole.objects.filter(pk=obj.project_role.id).update(is_active=True)
        return response


@login_required
def export_task_to_csv(request, project_id, role):
    tasks = ProjectTask.objects.filter(project_role=role)
    wb = Workbook()
    ws1 = wb.active
    ws1.append(['S/N', 'Description', 'Discipline', 'Role', 'Uniclass Code', 'Is_Active'])
    for projecttasks in tasks:
        ws1.append([
            '',
            projecttasks.description,
            projecttasks.discipline.__str__(),
            projecttasks.project_role.role.__str__(),
            projecttasks.uniclass_code,
            projecttasks.project_role.is_active
        ])

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

    response = HttpResponse(content=stream, content_type='application/ms-excel',)
    response['Content-Disposition'] = f'attachment; filename=Tasks-{project_id}.xlsx'
    return response


@login_required()
def import_roles(request):
    if request.method == 'POST':
        form = forms.Importer(request.POST, request.FILES)
        if form.is_valid():
            is_header = form.cleaned_data.get('header')
            excel_file = form.cleaned_data.get('excel_file')
            wb = load_workbook(excel_file)
            sheet = wb.active
            total = 0
            minrow = 2 if is_header else 1
            for row in sheet.iter_rows(min_row=minrow, values_only=True):
                if row[0] is not None and row[1] is not None:
                    obj, created = Role.objects.get_or_create(
                        name=row[1],
                        defaults={
                            "code":row[0]
                        }
                    )
                    if created:
                        total += 1

            messages.success(request, f'{total} item(s) saved to the db.')
            return redirect('tidp:roles')
        else:
            messages.warning(request, form.errors)

    form = forms.Importer()
    url = reverse_lazy('tidp:download-template', kwargs={'content_type':'role'})
    instruction = f'Please click <a href="{url}">here to download the template</a> . Empty rows will be ignored. If you want the first row to be included in the uploaded items, uncheck the "first row is header" button.'

    context = {
        "object": "roles",
        "form": form,
        "instruction": instruction,
    }
    return render(request, 'tidp/importer.html', context)


@login_required()
def import_tasktypes(request):
    if request.method == 'POST':
        form = forms.Importer(request.POST, request.FILES)
        if form.is_valid():
            is_header = form.cleaned_data.get('header')
            excel_file = form.cleaned_data.get('excel_file')
            wb = load_workbook(excel_file)
            sheet = wb.active
            total = 0
            minrow = 2 if is_header else 1
            for row in sheet.iter_rows(min_row=minrow, values_only=True):
                if row[0] is not None and row[1] is not None:
                    obj, created = TaskType.objects.get_or_create(
                        title=row[0],
                        defaults={"code":row[1]}
                    )
                    if created:
                        total +=1

            messages.success(request, f'{total} item(s) saved to the db.')
            return redirect('tidp:tasktypes')
        else:
            messages.warning(request, form.errors)

    form = forms.Importer()
    url = reverse_lazy('tidp:download-template', kwargs={'content_type':'tasktype'})
    instruction = f'Please click <a href="{url}">here to download the template</a> . Empty rows will be ignored. If you want the first row to be included in the uploaded items, uncheck the "first row is header" button.'

    context = {
        "object": "task types",
        "form": form,
        "instruction": instruction,
    }
    return render(request, 'tidp/importer.html', context)


@login_required
def import_tasks(request):
    errors = ''
    if request.method == 'POST':
        form = forms.ImportTaskForm(request.POST, request.FILES)
        if form.is_valid():
            is_header = form.cleaned_data.get('header')
            excel_file = form.cleaned_data.get('excel_file')
            discipline = form.cleaned_data.get('discipline')
            discipline_category = form.cleaned_data.get('discipline_category')
            wb = load_workbook(excel_file)
            sheet = wb.active
            total = 0
            minrow = 2 if is_header else 1
            for row in sheet.iter_rows(min_row=minrow, values_only=True):
                if row[0] is not None and row[1] is not None and row[2] is not None:
                    role = Role.objects.filter(name=row[0]).exists()
                    task_type = TaskType.objects.filter(title=row[1]).exists()
                    if role and task_type:
                        obj, created = Task.objects.get_or_create(
                            discipline=discipline,
                            discipline_category=discipline_category,
                            role=Role.objects.get(name=row[0]),
                            task_type=TaskType.objects.get(title=row[1]),
                            description=row[2],
                            defaults={
                                "uniclass_code":row[3] if row[3] else ''
                            }
                        )
                        if created:
                            total +=1
            messages.success(request, f'{total} tasks successfully added')
            return redirect('tidp:tasks')
        else:
            errors = form.errors

    form = forms.ImportTaskForm()
    description = f'Import an excel file that has the following columns respectively: "Role", "Task Type": "Task Title", "Uniclass code". Only uniclass code is optional. Rows not matching any roles or task type will be ignored'

    context = {
        "form": form,
        "errors": errors,
        "description": description,
    }
    return render(request, 'tidp/project/projecttask_import.html', context)


@login_required
def import_project_task(request, project_id, role):
    project = Project.objects.get(pk=project_id)
    role = ProjectRole.objects.get(pk=role)
    errors = ''
    if request.method == 'POST':
        form = forms.ImportProjectTaskForm(request.POST, request.FILES)
        if form.is_valid():
            is_header = form.cleaned_data.get('header')
            excel_file = form.cleaned_data.get('excel_file')
            wb = load_workbook(excel_file)
            sheet = wb.active
            total = 0
            minrow = 2 if is_header else 1
            for row in sheet.iter_rows(min_row=minrow, max_col=3, values_only=True):
                if row[0] is not None and row[1] is not None:
                    task_type = TaskType.objects.filter(title=row[0]).exists()
                    if task_type:
                        obj, created = ProjectTask.objects.get_or_create(
                            discipline=project.discipline,
                            discipline_category=project.discipline_category,
                            project_role=role,
                            task_type=TaskType.objects.get(title=row[0]),
                            description=row[1],
                            defaults={
                                "uniclass_code":row[2] if row[2] else ''
                            }
                        )
                        if created:
                            total += 1
            messages.success(request, f'{total} tasks successfully added')
            if request.user.is_superuser:
                url = 'tidp:project-detail'
            else:
                url = 'tidp:project-detail-leader'
            return redirect(url, project_id)


        else:
            errors = form.errors

    form = forms.ImportProjectTaskForm()
    description = f'Import an excel file that has the following columns respectively: "Task Type": "Task Title", "Uniclass code". Only uniclass code is optional. Rows not matching any roles or task type will be ignored'

    context = {
        "form": form,
        "errors": errors,
        "description": description,
    }
    return render(request, 'tidp/project/projecttask_import.html', context)


@login_required
def get_company(request):
    pk = request.GET.get('pk')
    company = Company.objects.get(pk=pk)
    data = {
        "contact_email": company.email,
        "contact_person": company.contact_person
    }
    return JsonResponse(data)


@login_required
def get_disciplinecategory(request):
    disciplines = DisciplineCategory.objects.filter(discipline_id=request.GET.get('pk'))
    data = []
    for dis in disciplines:
        data.append({"key": dis.id, "value": dis.__str__()})
    return JsonResponse(data, safe=False)


class CompleteTaskPage(LoginRequiredMixin, View):
    formset = forms.CompleteTasks
    template_name = "tidp/project/complete_tasks.html"

    def dispatch(self, request, *args, **kwargs):
        project = Project.objects.get(pk=kwargs['project_id'])
        projectrole = ProjectRole.objects.get(role_id=kwargs['role_id'], company=kwargs['company_id'], project=project)
        tasks = ProjectTask.objects.filter(project_role=projectrole).order_by('task_type')
        self.project = project
        self.tasks = tasks
        self.projectrole = projectrole
        form_class = f"{self.project.midp_tipd_template}_form"
        self.form = ''.join(x for x in form_class.title() if x != "_")

        return super(CompleteTaskPage, self).dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        formset = self.formset(queryset=self.tasks)
        formset.form = eval(f"forms.{self.form}")

        context = {
            "project": self.project,
            "formset": formset,
        }
        return render(request, self.template_name, context)


    def post(self, request, *args, **kwargs):
        formset = self.formset(request.POST, queryset=self.tasks)
        formset.form = eval(f"forms.{self.form}")
        if formset.is_valid():
            formset.save()
            grouped = {}
            for obj in self.tasks:
                grouped.setdefault(obj.task_type, []).append(obj)

            for values in grouped.values():
                for i, obj in enumerate(values, 1):
                    obj.code = i
                    obj.save(update_fields=['code'])

            self.projectrole.is_active = False
            self.projectrole.save()
            if request.user.is_superuser:
                url = 'tidp:project-detail'
            else:
                url = 'tidp:project-detail-leader'
            return redirect(url, self.project.id)

        context = {
            "project": self.project,
            "formset": formset,
        }
        return render(request, self.template_name, context)


def download_exl_template(request, content_type):
    wb = Workbook()
    ws1 = wb.active
    if content_type == 'role':
        ws1.append(['code', 'name'])
    elif content_type == 'tasktype':
        ws1.append(['title', 'code'])

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

    response = HttpResponse(content=stream, content_type='application/ms-excel',)
    filename = f'{content_type} Template'
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    return response


class GenerateTidpMidp(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        project = Project.objects.get(pk=kwargs['project_id'])
        tasks = ProjectTask.objects.filter(project_role__project=project, project_role__status='approved')
        template = settings.BASE_DIR / f"tidp/excel_templates/midp/{project.midp_tipd_template}.xlsx"
        filename = f'MIDP-{project.name}.xlsx'

        if kwargs.get('projectrole_id'):
            projectrole = ProjectRole.objects.get(pk=kwargs['projectrole_id'])
            tasks = ProjectTask.objects.filter(project_role_id=kwargs['projectrole_id'], project_role__status='approved')
            template = settings.BASE_DIR / f"tidp/excel_templates/tidp/{project.midp_tipd_template}.xlsx"
            filename = f'TIDP-{project.name}-{projectrole.role.name}.xlsx'

        wb = load_workbook(template)
        ws1 = wb.active

        #standard template
        if project.midp_tipd_template == 'standard_template':
            ws1['C4'].value = project.code
            ws1['C5'].value = project.name
            ws1['C6'].value = project.author.email
            ws1['C7'].value = project.created
            ws1['C8'].value = project.updated

            row = 16
            for task in tasks:
                ws1[f'B{row}'].value = task.description
                ws1[f'C{row}'].value = task.unique_code
                ws1[f'E{row}'].value = task.project_role.project.code
                ws1[f'F{row}'].value = task.project_role.project.client.originator_code
                ws1[f'G{row}'].value = task.volume
                ws1[f'H{row}'].value = task.level
                ws1[f'I{row}'].value = task.get_tasktype
                ws1[f'J{row}'].value = task.get_role
                ws1[f'K{row}'].value = format(task.code, '05d')
                ws1[f'L{row}'].value = task.uniclass_code

                row += 1

        #scottish template
        if project.midp_tipd_template == 'scottish_template':

            row = 14
            for task in tasks:
                ws1[f'C{row}'].value = task.uniclass_code
                ws1[f'D{row}'].value = task.description
                ws1[f'E{row}'].value = task.scale_or_size
                ws1[f'F{row}'].value = task.eir_reference
                ws1[f'G{row}'].value = task.pir_reference
                ws1[f'H{row}'].value = task.air_reference
                ws1[f'I{row}'].value = task.project_role.project.code
                ws1[f'J{row}'].value = task.project_role.project.client.originator_code
                ws1[f'K{row}'].value = task.volume
                ws1[f'L{row}'].value = task.level
                ws1[f'N{row}'].value = task.project_role.role.name
                ws1[f'O{row}'].value = task.format(task.code, '05d')

                row += 1


        #balfour_beatty_template template
        if project.midp_tipd_template == 'balfour_beatty_template':
            ws1['J6'].value = project.client.company_name
            ws1['J7'].value = project.name
            ws1['J8'].value = project.number
            ws1['J9'].value = project.discipline.name
            ws1['M9'].value = project.updated

            row = 12
            for task in tasks:
                ws1[f'C{row}'].value = task.project_role.project.code
                ws1[f'D{row}'].value = task.project_role.project.client.originator_code
                ws1[f'E{row}'].value = task.volume
                ws1[f'F{row}'].value = task.location
                ws1[f'G{row}'].value = task.get_tasktype
                ws1[f'H{row}'].value = task.get_role
                ws1[f'I{row}'].value = format(task.code, '05d')
                ws1[f'J{row}'].value = task.description
                ws1[f'K{row}'].value = task.description
                ws1[f'L{row}'].value = task.project_role.role.name
                ws1[f'M{row}'].value = task.estimated_production_duration
                ws1[f'N{row}'].value = task.uniclass_code

                row += 1


        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(content=stream, content_type='application/ms-excel',)
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response


class ViewTidpMidp(LoginRequiredMixin, View):
    template_name = 'tidp/view_tidp.html'

    def get(self, request, *args, **kwargs):
        project = Project.objects.get(pk=kwargs['project_id'])
        tasks = ProjectTask.objects.filter(project_role__project=project)
        title = f"TIDP | View MIDP | {project.name}"

        if kwargs.get('projectrole_id'):
            projectrole = ProjectRole.objects.get(pk=kwargs['projectrole_id'])
            title = f"TIDP | View TIDP | {projectrole}"
            tasks = ProjectTask.objects.filter(project_role_id=kwargs['projectrole_id'])

        context = {
            "title": title,
            "tasks": tasks,
            "templatename": project.midp_tipd_template,
        }
        return render(request, self.template_name, context)


class ProjectScaleView(LoginRequiredMixin, TemplateView):
    template_name = 'tidp/project/project_scales.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['projectscale'] = ProjectScale.objects.all()
        context["form"] = forms.ProjectScaleForm()
        return context

    def post(self, request, *args, **kwargs):
        form = forms.ProjectScaleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Scale successfully added.')
            return redirect('tidp:project-scales')
        return render(request, self.template_name, {"form":form})



class ProjectScaleUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = ProjectScale
    form_class = forms.ProjectScaleForm
    template_name = "tidp/update_form.html"
    success_url = reverse_lazy('tidp:project-scales')
    context_object_name = 'form'
    success_message = 'Project scale updated.'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['model'] = 'Project Scale'
        return context


class ProjectScaleDeleteView(LoginRequiredMixin, DeleteView):
    model = ProjectScale
    template_name = "tidp/delete_form.html"
    success_url = reverse_lazy('tidp:project-scales')
